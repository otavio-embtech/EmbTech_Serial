import os
import re
from collections import deque
from copy import deepcopy
from datetime import datetime

from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QCheckBox, QComboBox, QDialog, QDialogButtonBox, QFileDialog, QFormLayout,
    QGridLayout, QGroupBox, QHBoxLayout, QLabel, QLineEdit, QMessageBox,
    QPushButton, QScrollArea, QSpinBox, QTableWidget, QTableWidgetItem, QTextBrowser, QTextEdit,
    QVBoxLayout, QWidget,
)


PRINT_SENSORS_SAMPLE = """---SENSORES----
Atrib 0x5100: Entradas: 0x0
Atrib 0x5103: Camara: 27 - Vapor/RI: 27
Atrib 0x5104: Sonda Geral: 27 - Placa: 28
Atrib 0x5105: Sonda 1: 27 - Sonda 2: 27
Atrib 0x5106: Sonda 3: 27 - Sonda 4: 27
Atrib 0x5107: Sonda 5: 27 - Sonda 6: 27
Atrib 0x5110: AN1: 3mV
Atrib 0x5111: AN2: 3mV
Atrib 0x511A: Tensao AC: 40mV
Atrib 0x511B: IAC: 9mA
Atrib 0x5120: I-IHM: 103mA
Atrib 0x5125: Ult.Vol.: 0ml"""


PRESET_PROFILES = {
    "Generico": {
        "capture_mode": "event",
        "snapshot_window_ms": 300,
        "enabled_types": ["recebido", "enviado"],
        "sample_type": "recebido",
        "sample_port": "Principal",
        "sample_message": PRINT_SENSORS_SAMPLE,
        "rules": [
            {"enabled": True, "header": "Valor 1", "column": "F", "message_type": "recebido", "port_filter": "", "regex": r"([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Valor 2", "column": "G", "message_type": "recebido", "port_filter": "", "regex": r"[-+]?\d+(?:\.\d+)?\D+([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
        ],
    },
    "PRINT_SENSORS EmbTech": {
        "capture_mode": "snapshot",
        "snapshot_window_ms": 1200,
        "enabled_types": ["recebido"],
        "sample_type": "recebido",
        "sample_port": "Principal",
        "sample_message": PRINT_SENSORS_SAMPLE,
        "rules": [
            {"enabled": True, "header": "Entradas", "column": "F", "message_type": "recebido", "port_filter": "", "regex": r"Entradas:\s*(0x[0-9A-Fa-f]+)", "value_mode": "group1"},
            {"enabled": True, "header": "Camara_C", "column": "G", "message_type": "recebido", "port_filter": "", "regex": r"Camara:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Vapor_RI_C", "column": "H", "message_type": "recebido", "port_filter": "", "regex": r"Vapor/RI:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_Geral_C", "column": "I", "message_type": "recebido", "port_filter": "", "regex": r"Sonda Geral:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Placa_C", "column": "J", "message_type": "recebido", "port_filter": "", "regex": r"Placa:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_1_C", "column": "K", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 1:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_2_C", "column": "L", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 2:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_3_C", "column": "M", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 3:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_4_C", "column": "N", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 4:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_5_C", "column": "O", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 5:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Sonda_6_C", "column": "P", "message_type": "recebido", "port_filter": "", "regex": r"Sonda 6:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "AN1_mV", "column": "Q", "message_type": "recebido", "port_filter": "", "regex": r"AN1:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
            {"enabled": True, "header": "AN2_mV", "column": "R", "message_type": "recebido", "port_filter": "", "regex": r"AN2:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
            {"enabled": True, "header": "Tensao_AC_mV", "column": "S", "message_type": "recebido", "port_filter": "", "regex": r"Tensao AC:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
            {"enabled": True, "header": "IAC_mA", "column": "T", "message_type": "recebido", "port_filter": "", "regex": r"IAC:\s*([-+]?\d+(?:\.\d+)?)mA", "value_mode": "group1"},
            {"enabled": True, "header": "I_IHM_mA", "column": "U", "message_type": "recebido", "port_filter": "", "regex": r"I-IHM:\s*([-+]?\d+(?:\.\d+)?)mA", "value_mode": "group1"},
            {"enabled": True, "header": "Ult_Vol_ml", "column": "V", "message_type": "recebido", "port_filter": "", "regex": r"Ult\\.Vol\\.:\s*([-+]?\d+(?:\.\d+)?)ml", "value_mode": "group1"},
        ],
    },
    "Sensores EmbTech": {
        "capture_mode": "snapshot",
        "snapshot_window_ms": 1000,
        "enabled_types": ["recebido"],
        "sample_type": "recebido",
        "sample_port": "Principal",
        "sample_message": PRINT_SENSORS_SAMPLE,
        "rules": [
            {"enabled": True, "header": "Sonda_Geral_C", "column": "F", "message_type": "recebido", "port_filter": "", "regex": r"Sonda Geral:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "Placa_C", "column": "G", "message_type": "recebido", "port_filter": "", "regex": r"Placa:\s*([-+]?\d+(?:\.\d+)?)", "value_mode": "group1"},
            {"enabled": True, "header": "AN1_mV", "column": "H", "message_type": "recebido", "port_filter": "", "regex": r"AN1:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
            {"enabled": True, "header": "AN2_mV", "column": "I", "message_type": "recebido", "port_filter": "", "regex": r"AN2:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
            {"enabled": True, "header": "IAC_mA", "column": "J", "message_type": "recebido", "port_filter": "", "regex": r"IAC:\s*([-+]?\d+(?:\.\d+)?)mA", "value_mode": "group1"},
            {"enabled": True, "header": "I_IHM_mA", "column": "K", "message_type": "recebido", "port_filter": "", "regex": r"I-IHM:\s*([-+]?\d+(?:\.\d+)?)mA", "value_mode": "group1"},
            {"enabled": True, "header": "Tensao_AC_mV", "column": "L", "message_type": "recebido", "port_filter": "", "regex": r"Tensao AC:\s*([-+]?\d+(?:\.\d+)?)mV", "value_mode": "group1"},
        ],
    },
    "Estados Digitais": {
        "capture_mode": "snapshot",
        "snapshot_window_ms": 600,
        "enabled_types": ["recebido"],
        "sample_type": "recebido",
        "sample_port": "Principal",
        "sample_message": PRINT_SENSORS_SAMPLE,
        "rules": [
            {"enabled": True, "header": "Entradas", "column": "F", "message_type": "recebido", "port_filter": "", "regex": r"Entradas:\s*(0x[0-9A-Fa-f]+)", "value_mode": "group1"},
            {"enabled": True, "header": "Saidas", "column": "G", "message_type": "recebido", "port_filter": "", "regex": r"Saidas:\s*(0x[0-9A-Fa-f]+)", "value_mode": "group1"},
        ],
    },
}

PRESET_RULES = {name: deepcopy(profile["rules"]) for name, profile in PRESET_PROFILES.items()}


def build_default_datalogger_settings():
    return {
        "file_path": "",
        "sheet_name": "DataLogger",
        "header_row": 1,
        "enabled_types": ["recebido", "enviado"],
        "capture_mode": "event",
        "snapshot_window_ms": 300,
        "preview_history_size": 80,
        "preset_name": "Generico",
        "base_columns": {"timestamp": "A", "type": "B", "port": "C", "latency_ms": "D", "message": "E"},
        "rules": deepcopy(PRESET_RULES["Generico"]),
    }


def _normalize_column_letter(value):
    text = str(value or "").strip().upper()
    if text and not re.fullmatch(r"[A-Z]{1,3}", text):
        raise ValueError(f"Coluna invalida: {value}")
    return text


def _column_index(letter):
    index = 0
    for char in _normalize_column_letter(letter):
        index = index * 26 + (ord(char) - 64)
    return index


def _to_float(value):
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


class PreviewCanvas(FigureCanvas):
    def __init__(self):
        self.figure = Figure(figsize=(5.4, 2.1), constrained_layout=False)
        super().__init__(self.figure)
        self.axis = self.figure.add_subplot(111)
        self.setMinimumHeight(200)
        self.setMinimumWidth(260)
        self.draw_empty()

    def draw_empty(self, title="Preview"):
        self.figure.clear()
        self.axis = self.figure.add_subplot(111)
        self.axis.set_title(title, fontsize=10, fontweight="bold")
        self.axis.text(0.5, 0.5, "Sem dados", ha="center", va="center", transform=self.axis.transAxes)
        self.axis.set_xticks([])
        self.axis.set_yticks([])
        self.draw_idle()

    def draw_series(self, title, samples):
        if not samples:
            self.draw_empty(title)
            return
        self.figure.clear()
        self.axis = self.figure.add_subplot(111)
        self.axis.set_title(title, fontsize=10, fontweight="bold")
        x_vals = list(range(1, len(samples) + 1))
        y_vals = [item["value"] for item in samples]
        self.axis.plot(x_vals, y_vals, marker="o", color="#1f77b4", linewidth=1.7)
        self.axis.fill_between(x_vals, y_vals, color="#9EC5FE", alpha=0.35)
        self.axis.grid(axis="y", linestyle="--", alpha=0.25)
        self.axis.set_xlabel("Amostra")
        self.axis.set_ylabel("Valor")
        self.draw_idle()


class DataLoggerHelpDialog(QDialog):
    def __init__(self, markdown_text, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ajuda do DataLogger")
        self.resize(840, 680)
        layout = QVBoxLayout(self)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(True)
        browser.setMarkdown(markdown_text)
        layout.addWidget(browser)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(self.reject)
        buttons.accepted.connect(self.accept)
        buttons.button(QDialogButtonBox.StandardButton.Close).clicked.connect(self.accept)
        layout.addWidget(buttons)


class DataLoggerManager:
    BASE_HEADERS = {"timestamp": "DataHora", "type": "Tipo", "port": "Porta", "latency_ms": "Latencia_ms", "message": "Mensagem"}

    def __init__(self, settings=None):
        self.settings = build_default_datalogger_settings()
        self.preview_history = {}
        self._snapshot_state = {"row": None, "timestamp": None, "port": ""}
        self.update_settings(settings or {})

    def update_settings(self, settings):
        merged = build_default_datalogger_settings()
        if isinstance(settings, dict):
            merged.update(settings)
            merged["base_columns"].update(settings.get("base_columns", {}))
            if "rules" in settings:
                merged["rules"] = deepcopy(settings.get("rules") or [])
        self.settings = merged
        self._ensure_preview_buckets()

    def _ensure_preview_buckets(self):
        history_size = max(10, int(self.settings.get("preview_history_size", 80) or 80))
        current = dict(self.preview_history)
        self.preview_history = {}
        for rule in self.settings.get("rules", []):
            header = str(rule.get("header", "")).strip()
            if header:
                self.preview_history[header] = deque(current.get(header, []), maxlen=history_size)

    def get_preview_state(self):
        return {"history": {k: [dict(v) for v in values] for k, values in self.preview_history.items()}}

    def load_preview_state(self, state):
        history_size = max(10, int(self.settings.get("preview_history_size", 80) or 80))
        self.preview_history = {
            str(k): deque((dict(item) for item in values), maxlen=history_size)
            for k, values in (state or {}).get("history", {}).items()
        }
        self._ensure_preview_buckets()

    def ensure_workbook(self):
        path = str(self.settings.get("file_path", "")).strip()
        if not path:
            raise ValueError("Nenhum arquivo configurado para o DataLogger.")
        folder = os.path.dirname(path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        workbook = load_workbook(path) if os.path.exists(path) else Workbook()
        sheet_name = str(self.settings.get("sheet_name", "DataLogger")).strip() or "DataLogger"
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        worksheet.title = sheet_name
        header_row = max(1, int(self.settings.get("header_row", 1) or 1))
        for key, header in self.BASE_HEADERS.items():
            letter = str(self.settings.get("base_columns", {}).get(key, "")).strip()
            if letter:
                worksheet[f"{_normalize_column_letter(letter)}{header_row}"] = header
        for rule in self.settings.get("rules", []):
            if rule.get("enabled", True) and rule.get("header") and rule.get("column"):
                worksheet[f"{_normalize_column_letter(rule['column'])}{header_row}"] = str(rule["header"]).strip()
        workbook.save(path)

    def _event_enabled(self, event):
        enabled = set(self.settings.get("enabled_types") or [])
        return not enabled or str(event.get("msg_type", "")).strip().lower() in enabled

    def extract_values(self, event):
        msg = str(event.get("message", "") or "")
        port = str(event.get("source_port", "") or "")
        msg_type = str(event.get("msg_type", "") or "").lower()
        extracted = []
        for rule in self.settings.get("rules", []):
            if not rule.get("enabled", True):
                continue
            header = str(rule.get("header", "")).strip()
            column = str(rule.get("column", "")).strip()
            if not header or not column:
                continue
            rule_type = str(rule.get("message_type", "qualquer")).strip().lower()
            if rule_type not in ("", "qualquer", msg_type):
                continue
            port_filter = str(rule.get("port_filter", "")).strip()
            if port_filter and port_filter.lower() not in port.lower():
                continue
            value_mode = str(rule.get("value_mode", "group1")).strip().lower()
            regex = str(rule.get("regex", "") or "")
            value = ""
            if value_mode == "message":
                value = msg
            else:
                match = re.search(regex, msg, re.IGNORECASE) if regex else None
                if not match:
                    continue
                value = match.group(0) if value_mode == "match" else (match.group(1) if match.lastindex else match.group(0))
            extracted.append({"header": header, "column": _normalize_column_letter(column), "value": value, "numeric_value": _to_float(value)})
        return extracted

    def _append_preview(self, timestamp, extracted):
        self._ensure_preview_buckets()
        for item in extracted:
            if item["numeric_value"] is None:
                continue
            self.preview_history[item["header"]].append({"timestamp": timestamp.strftime("%H:%M:%S"), "value": item["numeric_value"]})

    def _resolve_row(self, worksheet, timestamp, port):
        header_row = max(1, int(self.settings.get("header_row", 1) or 1))
        if str(self.settings.get("capture_mode", "event")).strip().lower() != "snapshot":
            return max(header_row + 1, worksheet.max_row + 1)
        window_ms = max(50, int(self.settings.get("snapshot_window_ms", 300) or 300))
        if self._snapshot_state["row"] and self._snapshot_state["timestamp"] and self._snapshot_state["port"] == port:
            delta_ms = abs((timestamp - self._snapshot_state["timestamp"]).total_seconds() * 1000.0)
            if delta_ms <= window_ms:
                self._snapshot_state["timestamp"] = timestamp
                return self._snapshot_state["row"]
        row = max(header_row + 1, worksheet.max_row + 1)
        self._snapshot_state = {"row": row, "timestamp": timestamp, "port": port}
        return row

    def append_event(self, event):
        path = str(self.settings.get("file_path", "")).strip()
        if not path or not self._event_enabled(event):
            return []
        timestamp = event.get("timestamp", datetime.now())
        extracted = self.extract_values(event)
        self._append_preview(timestamp, extracted)
        self.ensure_workbook()
        workbook = load_workbook(path)
        worksheet = workbook[str(self.settings.get("sheet_name", "DataLogger")).strip() or "DataLogger"]
        row = self._resolve_row(worksheet, timestamp, str(event.get("source_port", "") or ""))
        base_columns = self.settings.get("base_columns", {})
        base_values = {
            "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
            "type": event.get("msg_type", ""),
            "port": event.get("source_port", ""),
            "latency_ms": event.get("latency_ms", ""),
            "message": event.get("message", ""),
        }
        for key, value in base_values.items():
            letter = str(base_columns.get(key, "")).strip()
            if letter:
                worksheet.cell(row=row, column=_column_index(letter), value=value)
        for item in extracted:
            worksheet.cell(row=row, column=_column_index(item["column"]), value=item["value"])
        workbook.save(path)
        return extracted


class DataLoggerConfigDialog(QDialog):
    MESSAGE_TYPES = [("qualquer", "Qualquer"), ("recebido", "Recebido"), ("enviado", "Enviado"), ("sistema", "Sistema"), ("erro", "Erro"), ("informacao", "Informacao")]
    VALUE_MODES = [("group1", "Grupo 1"), ("match", "Match completo"), ("message", "Mensagem completa")]

    def __init__(self, settings=None, preview_state=None, parent=None, embedded=False):
        super().__init__(parent)
        self._embedded = bool(embedded)
        self._embedded_save_handler = None
        self.embedded_save_button = None
        if self._embedded:
            self.setWindowFlags(Qt.WindowType.Widget)
        self.setWindowTitle("Configurar DataLogger")
        self.resize(940, 620)
        self.setMinimumSize(820, 520)
        self._settings = build_default_datalogger_settings()
        if settings:
            self._settings.update(settings)
            self._settings["base_columns"].update(settings.get("base_columns", {}))
            self._settings["rules"] = deepcopy(settings.get("rules", self._settings["rules"]))
        self._preview_manager = DataLoggerManager(self._settings)
        self._preview_manager.load_preview_state(preview_state or {})
        self._init_ui()
        self._load_settings()
        self._refresh_preview_rule_combo()
        self._refresh_preview_views([])

    def _get_preset_profile(self, preset_name):
        return PRESET_PROFILES.get(str(preset_name or "").strip(), {})

    def _set_enabled_type_checkboxes(self, enabled_types):
        enabled_set = set(enabled_types or [])
        for key, cb in self.type_checkboxes.items():
            cb.setChecked(key in enabled_set)

    def _load_sample_for_preset(self, preset_name, force=False):
        profile = self._get_preset_profile(preset_name)
        sample_message = str(profile.get("sample_message", "") or "")
        sample_type = str(profile.get("sample_type", "recebido") or "recebido")
        sample_port = str(profile.get("sample_port", "Principal") or "Principal")
        if force or not self.sample_message_input.toPlainText().strip():
            self.sample_message_input.setPlainText(sample_message)
        if force or not self.sample_port_input.text().strip():
            self.sample_port_input.setText(sample_port)
        sample_index = self.sample_type_combo.findData(sample_type)
        if sample_index >= 0:
            self.sample_type_combo.setCurrentIndex(sample_index)

    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(8)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QScrollArea.Shape.NoFrame)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(8)

        header_layout = QHBoxLayout()
        title_label = QLabel("Configuração avançada do DataLogger")
        title_label.setStyleSheet("font-size: 15px; font-weight: bold;")
        help_button = QPushButton("?")
        help_button.setFixedWidth(32)
        help_button.setToolTip("Abrir guia de uso do DataLogger")
        help_button.clicked.connect(self._open_help_dialog)
        header_layout.addWidget(title_label)
        header_layout.addStretch(1)
        header_layout.addWidget(help_button)
        content_layout.addLayout(header_layout)

        top = QGroupBox("Arquivo, preset e captura")
        top_layout = QGridLayout(top)
        self.file_path_input = QLineEdit()
        choose_button = QPushButton("Escolher arquivo...")
        choose_button.clicked.connect(self._choose_file)
        self.sheet_name_input = QLineEdit()
        self.header_row_spin = QSpinBox(); self.header_row_spin.setRange(1, 9999)
        self.capture_mode_combo = QComboBox(); self.capture_mode_combo.addItem("Evento por linha", "event"); self.capture_mode_combo.addItem("Snapshot consolidado", "snapshot")
        self.snapshot_window_spin = QSpinBox(); self.snapshot_window_spin.setRange(50, 10000); self.snapshot_window_spin.setSuffix(" ms")
        self.preset_combo = QComboBox(); self.preset_combo.addItem("Personalizado"); [self.preset_combo.addItem(name) for name in PRESET_RULES.keys()]
        preset_button = QPushButton("Aplicar preset"); preset_button.clicked.connect(self._apply_preset)
        self.preset_hint_label = QLabel("Dica: para PRINT_SENSORS use o preset 'PRINT_SENSORS EmbTech'.")
        self.preset_hint_label.setWordWrap(True)
        self.preset_hint_label.setStyleSheet("color: #C8C8C8; font-size: 11px;")
        top_layout.addWidget(QLabel("Arquivo:"), 0, 0); top_layout.addWidget(self.file_path_input, 0, 1, 1, 3); top_layout.addWidget(choose_button, 0, 4)
        top_layout.addWidget(QLabel("Planilha:"), 1, 0); top_layout.addWidget(self.sheet_name_input, 1, 1); top_layout.addWidget(QLabel("Cabecalho:"), 1, 2); top_layout.addWidget(self.header_row_spin, 1, 3)
        top_layout.addWidget(QLabel("Modo:"), 2, 0); top_layout.addWidget(self.capture_mode_combo, 2, 1); top_layout.addWidget(QLabel("Janela snapshot:"), 2, 2); top_layout.addWidget(self.snapshot_window_spin, 2, 3)
        top_layout.addWidget(QLabel("Preset:"), 3, 0); top_layout.addWidget(self.preset_combo, 3, 1, 1, 2); top_layout.addWidget(preset_button, 3, 3)
        top_layout.addWidget(self.preset_hint_label, 4, 0, 1, 5)
        content_layout.addWidget(top)

        types = QGroupBox("Tipos de evento")
        types_layout = QHBoxLayout(types)
        self.type_checkboxes = {}
        for key, label in self.MESSAGE_TYPES[1:]:
            cb = QCheckBox(label); self.type_checkboxes[key] = cb; types_layout.addWidget(cb)
        types_layout.addStretch(1)
        content_layout.addWidget(types)

        middle = QHBoxLayout()
        base_group = QGroupBox("Colunas base")
        base_layout = QFormLayout(base_group)
        self.base_column_inputs = {}
        for key, label in (("timestamp", "Data/Hora"), ("type", "Tipo"), ("port", "Porta"), ("latency_ms", "Latencia"), ("message", "Mensagem")):
            edit = QLineEdit(); edit.setPlaceholderText("Ex.: A"); self.base_column_inputs[key] = edit; base_layout.addRow(label + ":", edit)
        middle.addWidget(base_group, 1)

        rules_group = QGroupBox("Regras")
        rules_layout = QVBoxLayout(rules_group)
        self.rules_table = QTableWidget(0, 7)
        self.rules_table.setHorizontalHeaderLabels(["Ativo", "Cabecalho", "Coluna", "Tipo", "Porta", "Regex", "Valor"])
        self.rules_table.horizontalHeader().setStretchLastSection(True)
        rules_layout.addWidget(self.rules_table)
        row_buttons = QHBoxLayout()
        add_button = QPushButton("Adicionar regra"); add_button.clicked.connect(self._add_rule_row)
        remove_button = QPushButton("Remover regra"); remove_button.clicked.connect(self._remove_selected_rule)
        row_buttons.addWidget(add_button); row_buttons.addWidget(remove_button); row_buttons.addStretch(1)
        rules_layout.addLayout(row_buttons)
        middle.addWidget(rules_group, 2)
        content_layout.addLayout(middle, 1)

        preview = QGroupBox("Monitor e preview")
        preview_layout = QGridLayout(preview)
        self.preview_rule_combo = QComboBox(); self.preview_rule_combo.currentIndexChanged.connect(self._refresh_preview_chart)
        self.preview_rule_combo_2 = QComboBox(); self.preview_rule_combo_2.currentIndexChanged.connect(self._refresh_preview_chart_2)
        self.preview_rule_combo_3 = QComboBox(); self.preview_rule_combo_3.currentIndexChanged.connect(self._refresh_preview_chart_3)
        self.preview_rule_combos = [self.preview_rule_combo, self.preview_rule_combo_2, self.preview_rule_combo_3]
        self.preview_canvas = PreviewCanvas()
        self.preview_canvas_2 = PreviewCanvas()
        self.preview_canvas_3 = PreviewCanvas()
        self.preview_canvases = [self.preview_canvas, self.preview_canvas_2, self.preview_canvas_3]
        self.preview_text = QTextEdit(); self.preview_text.setReadOnly(True); self.preview_text.setMaximumHeight(88)
        self.sample_type_combo = QComboBox(); [self.sample_type_combo.addItem(label, key) for key, label in self.MESSAGE_TYPES]
        self.sample_port_input = QLineEdit(); self.sample_port_input.setPlaceholderText("Principal / Modbus")
        self.sample_message_input = QTextEdit(); self.sample_message_input.setMaximumHeight(72)
        simulate_button = QPushButton("Simular no preview"); simulate_button.clicked.connect(self._simulate_preview)
        clear_button = QPushButton("Limpar preview"); clear_button.clicked.connect(self._clear_preview)
        preview_layout.addWidget(QLabel("Grafico 1:"), 0, 0); preview_layout.addWidget(self.preview_rule_combo, 0, 1)
        preview_layout.addWidget(QLabel("Grafico 2:"), 0, 2); preview_layout.addWidget(self.preview_rule_combo_2, 0, 3)
        preview_layout.addWidget(QLabel("Grafico 3:"), 0, 4); preview_layout.addWidget(self.preview_rule_combo_3, 0, 5)
        preview_layout.addWidget(clear_button, 0, 6)
        preview_layout.addWidget(self.preview_canvas, 1, 0, 1, 2)
        preview_layout.addWidget(self.preview_canvas_2, 1, 2, 1, 2)
        preview_layout.addWidget(self.preview_canvas_3, 1, 4, 1, 3)
        preview_layout.addWidget(QLabel("Ultima extração:"), 2, 0); preview_layout.addWidget(self.preview_text, 3, 0, 1, 7)
        preview_layout.addWidget(QLabel("Tipo:"), 4, 0); preview_layout.addWidget(self.sample_type_combo, 4, 1)
        preview_layout.addWidget(QLabel("Porta:"), 4, 2); preview_layout.addWidget(self.sample_port_input, 4, 3, 1, 4)
        preview_layout.addWidget(QLabel("Mensagem:"), 5, 0); preview_layout.addWidget(self.sample_message_input, 6, 0, 1, 7)
        preview_layout.addWidget(simulate_button, 7, 5, 1, 2)
        content_layout.addWidget(preview, 1)

        scroll_area.setWidget(content)
        root.addWidget(scroll_area, 1)

        if self._embedded:
            actions_layout = QHBoxLayout()
            actions_layout.addStretch(1)
            self.embedded_save_button = QPushButton("Salvar Configuração do DataLogger")
            self.embedded_save_button.clicked.connect(self._save_embedded_settings)
            actions_layout.addWidget(self.embedded_save_button)
            root.addLayout(actions_layout)
        else:
            box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
            box.accepted.connect(self.accept); box.rejected.connect(self.reject)
            root.addWidget(box)

    def _choose_file(self):
        default_name = datetime.now().strftime("EmbTech_DataLogger_%Y%m%d_%H%M%S.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Selecionar arquivo do DataLogger", self.file_path_input.text().strip() or default_name, "Planilhas Excel (*.xlsx)")
        if path:
            self.file_path_input.setText(path if path.lower().endswith(".xlsx") else path + ".xlsx")

    def _load_help_markdown(self):
        help_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "datalogger_help.md")
        if os.path.exists(help_path):
            with open(help_path, "r", encoding="utf-8") as handle:
                return handle.read()
        return "# Ajuda do DataLogger\n\nArquivo de ajuda não encontrado."

    def _open_help_dialog(self):
        try:
            dialog = DataLoggerHelpDialog(self._load_help_markdown(), self)
            dialog.exec()
        except Exception as exc:
            QMessageBox.warning(self, "Ajuda", f"Não foi possível abrir a ajuda do DataLogger:\n{exc}")

    def _add_rule_row(self, rule=None):
        row = self.rules_table.rowCount(); self.rules_table.insertRow(row); rule = rule or {}
        cb = QCheckBox(); cb.setChecked(bool(rule.get("enabled", True)))
        wrap = QWidget(); wrap_layout = QHBoxLayout(wrap); wrap_layout.setContentsMargins(0, 0, 0, 0); wrap_layout.addWidget(cb, alignment=Qt.AlignmentFlag.AlignCenter)
        self.rules_table.setCellWidget(row, 0, wrap)
        self.rules_table.setItem(row, 1, QTableWidgetItem(str(rule.get("header", ""))))
        self.rules_table.setItem(row, 2, QTableWidgetItem(str(rule.get("column", ""))))
        type_combo = QComboBox(); [type_combo.addItem(label, key) for key, label in self.MESSAGE_TYPES]; type_combo.setCurrentIndex(max(0, type_combo.findData(rule.get("message_type", "qualquer"))))
        value_combo = QComboBox(); [value_combo.addItem(label, key) for key, label in self.VALUE_MODES]; value_combo.setCurrentIndex(max(0, value_combo.findData(rule.get("value_mode", "group1"))))
        self.rules_table.setCellWidget(row, 3, type_combo); self.rules_table.setItem(row, 4, QTableWidgetItem(str(rule.get("port_filter", ""))))
        self.rules_table.setItem(row, 5, QTableWidgetItem(str(rule.get("regex", "")))); self.rules_table.setCellWidget(row, 6, value_combo)

    def _remove_selected_rule(self):
        if self.rules_table.currentRow() >= 0:
            self.rules_table.removeRow(self.rules_table.currentRow())
            self._refresh_preview_rule_combo()

    def _apply_preset(self):
        preset_name = self.preset_combo.currentText().strip()
        profile = self._get_preset_profile(preset_name)
        if not profile:
            return
        self.rules_table.setRowCount(0)
        for rule in profile.get("rules", []):
            self._add_rule_row(rule)
        capture_mode = str(profile.get("capture_mode", "event") or "event")
        capture_index = self.capture_mode_combo.findData(capture_mode)
        if capture_index >= 0:
            self.capture_mode_combo.setCurrentIndex(capture_index)
        self.snapshot_window_spin.setValue(int(profile.get("snapshot_window_ms", 300) or 300))
        self._set_enabled_type_checkboxes(profile.get("enabled_types", ["recebido"]))
        self._load_sample_for_preset(preset_name, force=True)
        self._refresh_preview_rule_combo()
        self._refresh_preview_views([])
        self._simulate_preview(silent=True)

    def _load_settings(self):
        self.file_path_input.setText(str(self._settings.get("file_path", "")))
        self.sheet_name_input.setText(str(self._settings.get("sheet_name", "DataLogger")))
        self.header_row_spin.setValue(int(self._settings.get("header_row", 1) or 1))
        self.capture_mode_combo.setCurrentIndex(max(0, self.capture_mode_combo.findData(self._settings.get("capture_mode", "event"))))
        self.snapshot_window_spin.setValue(int(self._settings.get("snapshot_window_ms", 300) or 300))
        preset_index = self.preset_combo.findText(str(self._settings.get("preset_name", "Personalizado"))); self.preset_combo.setCurrentIndex(preset_index if preset_index >= 0 else 0)
        enabled_types = set(self._settings.get("enabled_types") or [])
        for key, cb in self.type_checkboxes.items():
            cb.setChecked(key in enabled_types)
        for key, edit in self.base_column_inputs.items():
            edit.setText(str(self._settings.get("base_columns", {}).get(key, "")))
        self.rules_table.setRowCount(0)
        for rule in self._settings.get("rules", []):
            self._add_rule_row(rule)
        self._load_sample_for_preset(self.preset_combo.currentText().strip(), force=False)

    def _collect_rules(self):
        rules = []
        for row in range(self.rules_table.rowCount()):
            enabled_wrap = self.rules_table.cellWidget(row, 0); enabled_cb = enabled_wrap.findChild(QCheckBox) if enabled_wrap else None
            header = self.rules_table.item(row, 1).text().strip() if self.rules_table.item(row, 1) else ""
            column = self.rules_table.item(row, 2).text().strip() if self.rules_table.item(row, 2) else ""
            if not header and not column:
                continue
            if not header or not column:
                raise ValueError(f"Regra da linha {row + 1} incompleta.")
            type_combo = self.rules_table.cellWidget(row, 3); value_combo = self.rules_table.cellWidget(row, 6)
            rule = {
                "enabled": bool(enabled_cb.isChecked()) if enabled_cb else True,
                "header": header,
                "column": _normalize_column_letter(column),
                "message_type": type_combo.currentData() if type_combo else "qualquer",
                "port_filter": self.rules_table.item(row, 4).text().strip() if self.rules_table.item(row, 4) else "",
                "regex": self.rules_table.item(row, 5).text().strip() if self.rules_table.item(row, 5) else "",
                "value_mode": value_combo.currentData() if value_combo else "group1",
            }
            if rule["value_mode"] != "message" and not rule["regex"]:
                raise ValueError(f"Regra '{header}' precisa de regex.")
            rules.append(rule)
        return rules

    def _collect_preview_settings(self):
        settings = build_default_datalogger_settings()
        settings["capture_mode"] = self.capture_mode_combo.currentData() or "event"
        settings["snapshot_window_ms"] = self.snapshot_window_spin.value()
        settings["preview_history_size"] = self._settings.get("preview_history_size", 80)
        settings["rules"] = self._collect_rules()
        return settings

    def _refresh_preview_rule_combo(self):
        current_values = [combo.currentData() for combo in self.preview_rule_combos]
        headers = [str(rule["header"]) for rule in self._collect_preview_settings().get("rules", []) if rule.get("header")]
        for combo in self.preview_rule_combos:
            combo.blockSignals(True)
            combo.clear()
            for header in headers:
                combo.addItem(header, header)
            combo.blockSignals(False)

        for idx, combo in enumerate(self.preview_rule_combos):
            if not combo.count():
                continue
            current = current_values[idx]
            target_index = combo.findData(current)
            if target_index < 0:
                target_index = min(idx, combo.count() - 1)
            combo.setCurrentIndex(target_index)

        self._refresh_all_preview_charts()

    def _refresh_preview_chart(self):
        header = self.preview_rule_combo.currentData()
        self.preview_canvas.draw_series(header or "Preview", list(self._preview_manager.preview_history.get(header, [])) if header else [])

    def _refresh_preview_chart_2(self):
        header = self.preview_rule_combo_2.currentData()
        self.preview_canvas_2.draw_series(header or "Preview 2", list(self._preview_manager.preview_history.get(header, [])) if header else [])

    def _refresh_preview_chart_3(self):
        header = self.preview_rule_combo_3.currentData()
        self.preview_canvas_3.draw_series(header or "Preview 3", list(self._preview_manager.preview_history.get(header, [])) if header else [])

    def _refresh_all_preview_charts(self):
        self._refresh_preview_chart()
        self._refresh_preview_chart_2()
        self._refresh_preview_chart_3()

    def _refresh_preview_views(self, extracted):
        self._refresh_all_preview_charts()
        self.preview_text.setPlainText("\n".join(f"{item['header']} ({item['column']}): {item['value']}" for item in extracted) or "Nenhuma extração recente.")

    def _simulate_preview(self, silent=False):
        settings = self._collect_preview_settings()
        self._preview_manager.update_settings(settings)
        event = {
            "timestamp": datetime.now(),
            "msg_type": self.sample_type_combo.currentData() or "recebido",
            "source_port": self.sample_port_input.text().strip(),
            "message": self.sample_message_input.toPlainText().strip(),
            "latency_ms": 0,
        }
        extracted = self._preview_manager.extract_values(event)
        self._preview_manager._append_preview(event["timestamp"], extracted)
        self._refresh_preview_rule_combo()
        self._refresh_preview_views(extracted)
        if not extracted and not silent:
            QMessageBox.information(self, "Preview", "Nenhum valor foi extraido com as regras atuais.")

    def _clear_preview(self):
        self._preview_manager = DataLoggerManager(self._collect_preview_settings())
        self._refresh_preview_rule_combo()
        self._refresh_preview_views([])

    def consume_live_event(self, event):
        if not isinstance(event, dict):
            return []
        try:
            settings = self._collect_preview_settings()
            self._preview_manager.update_settings(settings)
            extracted = self._preview_manager.extract_values(event)
            self._preview_manager._append_preview(event.get("timestamp", datetime.now()), extracted)
            if self.preview_rule_combo.count() == 0:
                self._refresh_preview_rule_combo()
            self._refresh_preview_views(extracted)
            return extracted
        except Exception:
            return []

    def get_settings(self):
        file_path = self.file_path_input.text().strip()
        if not file_path:
            raise ValueError("Escolha um arquivo para o DataLogger.")
        settings = build_default_datalogger_settings()
        settings["file_path"] = file_path if file_path.lower().endswith(".xlsx") else file_path + ".xlsx"
        settings["sheet_name"] = self.sheet_name_input.text().strip() or "DataLogger"
        settings["header_row"] = self.header_row_spin.value()
        settings["enabled_types"] = [key for key, cb in self.type_checkboxes.items() if cb.isChecked()]
        settings["capture_mode"] = self.capture_mode_combo.currentData() or "event"
        settings["snapshot_window_ms"] = self.snapshot_window_spin.value()
        settings["preset_name"] = self.preset_combo.currentText().strip() or "Personalizado"
        settings["rules"] = self._collect_rules()
        for key, edit in self.base_column_inputs.items():
            settings["base_columns"][key] = _normalize_column_letter(edit.text().strip()) if edit.text().strip() else ""
        return settings

    def save_settings(self):
        self._settings = self.get_settings()
        return deepcopy(self._settings)

    def _save_embedded_settings(self):
        try:
            self.save_settings()
            if callable(self._embedded_save_handler):
                self._embedded_save_handler(deepcopy(self._settings))
            else:
                QMessageBox.information(self, "DataLogger", "Configuração salva.")
        except ValueError as exc:
            QMessageBox.warning(self, "DataLogger", str(exc))
        except Exception as exc:
            QMessageBox.warning(self, "DataLogger", f"Falha ao salvar a configuração:\n{exc}")

    def accept(self):
        try:
            self.save_settings()
        except ValueError as exc:
            QMessageBox.warning(self, "DataLogger", str(exc))
            return
        super().accept()

    def get_dialog_settings(self):
        return deepcopy(self._settings)
